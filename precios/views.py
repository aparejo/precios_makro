from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import Producto, Sucursal, ProductoEscaneado
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage
from .forms import SedeForm
from .forms import AdminForm, BarcodeForm, VerificadorForm, FondoVerificador
from precios.models import Usuario, Pantalla, BCV, Combo, Oferta, Roles, ProductoEscaneado
from .forms import PantallaForm, Pantalla, RolForm, ProductoForm
from precios.models import TareaActualizacion
from django.utils import timezone
from django.db.models import F
from datetime import datetime, date
from django.core.exceptions import ValidationError
from .models import Verificador, FondoVerificador
from django.utils.text import slugify
#from .utils import obtener_descripcion_limite
from gemini_utils import obtener_descripcion_limite
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from django.db.models import Count


from decimal import Decimal, ROUND_DOWN
from django.http import JsonResponse

@login_required
def lista_productos_escaneados(request):
    productos_escaneados = ProductoEscaneado.objects.all()
    sucursales = Sucursal.objects.all()
# Aplicar filtros si se envían parámetros en la solicitud
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')

    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        productos_escaneados = productos_escaneados.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        productos_escaneados = productos_escaneados.filter(fecha_hora__date__lte=fecha_fin)
    if sucursal_id:
        productos_escaneados = productos_escaneados.filter(sucursal_id=sucursal_id)

    # Calcular el contador de productos escaneados por sucursal del día de hoy
    hoy = date.today()
    productos_escaneados_hoy = productos_escaneados.filter(fecha_hora__date=hoy)
    contador_hoy = productos_escaneados_hoy.values('sucursal').annotate(contador=Count('id'))

    # Totalizar la cantidad de productos escaneados
    total_escaneadas = productos_escaneados.count()

    # Exportar a Excel si se envía el parámetro "export" en la solicitud
    export_to_excel = request.GET.get('export') == 'excel'
    if export_to_excel:
        # Crear una nueva consulta filtrada para la exportación a Excel
        productos_escaneados_export = productos_escaneados.values_list('codigo_producto', 'sucursal', 'precio', 'fecha_hora')

        # Crear un libro de Excel y una hoja de cálculo
        wb = Workbook()
        ws = wb.active

        # Agregar encabezados de columna
        headers = ['Código', 'Sucursal', 'Precio', 'Fecha y Hora']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws.column_dimensions[col_letter].width = 15

        # Agregar datos de productos escaneados filtrados
        for row_num, producto_escaneado in enumerate(productos_escaneados_export, 2):
            ws.cell(row=row_num, column=1, value=producto_escaneado[0])
            ws.cell(row=row_num, column=2, value=str(producto_escaneado[1]))
            ws.cell(row=row_num, column=3, value=str(producto_escaneado[2]))
            ws.cell(row=row_num, column=4, value=str(producto_escaneado[3]))

        # Crear una respuesta HTTP con el archivo de Excel adjunto
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=productos_escaneados.xlsx'
        wb.save(response)
        return response

    context = {
        'productos_escaneados': productos_escaneados,
        'fecha_inicio_filtro': fecha_inicio,
        'fecha_fin_filtro': fecha_fin,
        'sucursal_filtro': sucursal_id,
        'total_escaneadas': total_escaneadas,
        'contador_hoy': contador_hoy,
        'sucursales': sucursales  # Agrega la lista de sucursales al contexto
    }

    return render(request, 'lista_productos_escaneados.html', context)

@login_required
def editar_producto(request, codigo):
    producto = get_object_or_404(Producto, codigo=codigo)

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('productos')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'editar_producto.html', {'form': form})

@login_required
def crear_rol(request):
    if request.method == 'POST':
        form = RolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('administradores')
    else:
        form = RolForm()
    
    return render(request, 'crear_rol.html', {'form': form})

@login_required
def crear_verificador(request):
    if request.method == 'POST':
        form = VerificadorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('verificadores')  # Redireccionar a la página de verificadores
    else:
        form = VerificadorForm()

    context = {'form': form}
    return render(request, 'verificador/agregar_verificador.html', context)

@login_required
def verificador_detail(request, slug):
    verificador = get_object_or_404(Verificador, slug=slug)
    context = {'verificador': verificador}
    return render(request, 'verificador/agregar_verificador.html', context)

@login_required
def fondos_verificador(request):
    fondos = FondoVerificador.objects.all()
    context = {'fondos': fondos}
    return render(request, 'verificador/fondos.html', context)

def obtener_detalles_producto(request, codigo_barras):
    producto = Producto.objects.get(barra=codigo_barras)
    detalles = {
        'nombre': producto.descripcion,
        'codigo_producto': producto.codigo,
    }
    return JsonResponse(detalles)

def leer_codigo_de_barrascelular(request):
    context = {}
    ofertas = Oferta.objects.all()[:10] 
    
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form

    return render(request, 'validador_precios_vina.html', {**context, 'ofertas': ofertas})

def leer_codigo_de_barrasfinal(request, id):
    context = {}
    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[:20]
    sucursal = Sucursal.objects.get(codigo=id)
    precio = sucursal.slug
    
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto=producto)

                print(f'Combos encontrados: {combos}')

                pvp_base = getattr(producto, f'{precio}_pvp') * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = getattr(producto, f'{precio}_pvp')
                context['mensaje'] = 'Verificador de precios'
                
                producto_escaneado = ProductoEscaneado(
                    sucursal=sucursal,
                    codigo_producto=producto.codigo,
                    precio=pvp_base_bcv,
                    fecha_hora=timezone.now())
                producto_escaneado.save()
                
                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form

    return render(request, 'validador_precios_vina2.html', {**context, 'ofertas': ofertas})


def leer_codigo_de_barras(request):
    context = {}
    ofertas = Oferta.objects.all()[:10] 
    
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', {**context, 'ofertas': ofertas})

def leer_codigo_de_barrasT01(request):
    context = {}
    ofertas = Oferta.objects.all()[:10] 
    
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', {**context, 'ofertas': ofertas})

def leer_codigo_de_barrasT05(request):
    context = {}
    ofertas = Oferta.objects.all()[:10] 
    
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', {**context, 'ofertas': ofertas})

def leer_codigo_de_barrasV02(request):
    context = {}
    ofertas = Oferta.objects.all()[:10] 
    
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', {**context, 'ofertas': ofertas})

def leer_codigo_de_barrasT30(request):
    context = {}
    ofertas = Oferta.objects.all()[:10] 
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', {**context, 'ofertas': ofertas})

def leer_codigo_de_barrasT03(request):
    context = {}
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', context)

def leer_codigo_de_barrasT08(request):
    context = {}
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form

    return render(request, 'validador_precios_vina2.html', context)

def leer_codigo_de_barrasT25(request):
    context = {}
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', context)

def leer_codigo_de_barrasV21(request):
   context = {}
   if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
   else:
        form = BarcodeForm()
        context['form'] = form
   return render(request, 'validador_precios_vina2.html', context)

def leer_codigo_de_barrasT24(request):
    context = {}
    if request.method == 'POST':
        form = BarcodeForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                producto = Producto.objects.get(barra=barcode)
                precio_bcv = BCV.objects.latest('id').precio

                print(f'Producto encontrado: {producto}')

                combos = Combo.objects.filter(codigo_producto__iexact=producto.barra)

                print(f'Combos encontrados: {combos}')

                pvp_base = producto.La_Vina_pvp * Decimal(precio_bcv)
                pvp_base = pvp_base.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
                pvp_base_bcv = round(pvp_base, 2)

                context['producto'] = producto
                context['pvp_base_bcv'] = pvp_base_bcv
                context['pvp_base'] = producto.La_Vina_pvp

                if combos.exists():
                    combo = combos.first()
                    if combo.fecha_expiracion >= timezone.now().date():
                        context['combo'] = combo
                        context['descripcion_combo'] = combo.descripcion
                else:
                    print('No se encontraron combos para el producto.')
            except Producto.DoesNotExist:
                context['error'] = 'El código de barras no está asociado a ningún producto.'
            except BCV.DoesNotExist:
                context['error'] = 'No se encontró el precio del BCV.'
    else:
        form = BarcodeForm()
        context['form'] = form
    return render(request, 'validador_precios_vina2.html', context)

def mostrar_pantalla_mcy(request, nombre_pantalla):
    pantalla = Pantalla.objects.filter(nombre=nombre_pantalla).first()

    context = {
        'pantalla': pantalla,
    }
    return render(request, 'monitor-mcy.html', context)

@login_required
def pantalla_view(request):
    pantallas = Pantalla.objects.all()
    
    return render(request, 'pantallas.html', {'pantallas': pantallas})

@login_required
def ver_pantalla(request, pantalla_id):
    pantalla = get_object_or_404(Pantalla, id=pantalla_id)
    producto = pantalla.id_producto
    return render(request, 'ver_pantalla.html', {'pantalla': pantalla, 'producto': producto})

@login_required
def modificar_pantalla(request, pantalla_id):
    pantalla = get_object_or_404(Pantalla, id=pantalla_id)
    if request.method == 'POST':
        form = PantallaForm(request.POST, instance=pantalla)
        if form.is_valid():
            form.save()
            return redirect('ver_pantalla', pantalla_id=pantalla_id)
    else:
        form = PantallaForm(instance=pantalla)
    return render(request, 'modificar_pantalla.html', {'form': form, 'pantalla': pantalla})

@login_required
def ofertas_view(request):
    today = datetime.today()
    productos_oferta = Producto.objects.filter(la_Urbina_pvp__lt=F('pvp_base'))
    productos_oferta2 = Producto.objects.filter(la_Urbina_pvp__lt=F('pvp_base')).count()
    cantidad_medicinas = Oferta.objects.filter(linea__in=["MEDICAMENTOS", "VMS"]).count()
    cantidad_insumos = Oferta.objects.filter(linea__in=["INSUMOS MQ", "LABORATORIO", "BIODYNAMICS" ]).count()
    cantidad_equipos = Oferta.objects.filter(linea="EQUIPOS MEDICOS" ).count()
    cantidad_belleza = Oferta.objects.filter(linea__in=["COSMETICOS Y BELLEZA", "DERMATOLOGIA Y ESTETICA", "CUIDADO E HIGIENE PERSONAL" ]).count()
    cantidad_alimentos = Oferta.objects.filter(linea__in=["PANADERIA", "CHARCUTERÍA", "DAIRY", "CARNES ROJAS", "VIVERES", "BEBIDAS", "FRUTAS Y VEGETALES", "PESCADERIA", "CONGELADOS", "BASICOS"]).count()
    combos = Combo.objects.filter(fecha_expiracion__gte=today)  # Filtrar combos con fecha de expiración mayor o igual a la fecha actual
    context = {
        'Medicamentos': cantidad_medicinas,
        'Alimentos': cantidad_alimentos,
        'Bellezas': cantidad_belleza,
        'Equipos': cantidad_equipos,
        'Insumos': cantidad_insumos,
        'Ofertas2': productos_oferta2        
    }
    return render(request, 'precios_oferta.html', {'productos_oferta': productos_oferta, 'combos': combos, 'context': context})


def eliminar_pantalla(request, pantalla_id):
    pantalla = get_object_or_404(Pantalla, id=pantalla_id)
    if request.method == 'POST':
        # Eliminar la pantalla de la base de datos
        pantalla.delete()
        return redirect('lista_pantallas')
    else:
        return render(request, 'eliminar_pantalla.html', {'pantalla': pantalla})

@login_required
def agregar_pantalla(request):
    if request.method == 'POST':
        form = PantallaForm(request.POST)
        if form.is_valid():
            pantalla = form.save(commit=False)
            pantalla.save()
            return redirect('pantallas')  # Redirigir a la página de pantallas
    else:
        form = PantallaForm()
    return render(request, 'agregar_pantalla.html', {'form': form})

@login_required
def administradores(request):
    administradores = Usuario.objects.filter(is_superuser=False)
    return render(request, 'administradores.html', {'administradores': administradores})

@login_required
def agregar_administrador(request):
    if request.method == 'POST':
        form = AdminForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            sucursal = form.cleaned_data['sucursal']
            print(sucursal)
            usuario.sucursal = sucursal
            usuario.save()
            return redirect('administradores')
    else:
        form = AdminForm()

    return render(request, 'agregar_administrador.html', {'form': form})

@login_required
def BASE(request):
    bcv = BCV.objects.order_by('-fecha').first()  # Obtener el último registro
    precio_bcv = bcv.precio if bcv else None  # Obtener el precio o None si no hay registros
    fecha_bcv = bcv.fecha.strftime("%d/%m/%Y") if bcv else None  # Obtener la fecha formateada o None si no hay registros
    cantidad_ofertas = Producto.objects.filter(la_Urbina_pvp__lt=F('pvp_base')).count()
    cantidad_medicinas = Oferta.objects.filter(linea="MEDICAMENTOS").count()

    # Obtener las últimas 5 tareas de actualización
    ultimas_tareas = TareaActualizacion.objects.order_by('-fecha_actualizacion')[:5]

    # Obtener la hora actual del servidor
    hora_actual = timezone.now().strftime("%H:%M:%S")
    
    context = {
        'precio_bcv': precio_bcv,
        'fecha_bcv': fecha_bcv,
        'ultimas_tareas': ultimas_tareas,
        'Ofertas': cantidad_ofertas,
        'Medicamentos': cantidad_medicinas,
        'hora_actual': hora_actual  # Agregar la hora actual al contexto
    }
    return render(request, 'base.html', context)

@login_required
def productos(request):
    query = request.GET.get('q')
    productos_list = Producto.objects.filter(existencia_total__gt=0)  # Filtrar por existencia_total > 0

    if query:
        productos_list = productos_list.filter(
            Q(descripcion__icontains=query) |
            Q(barra__icontains=query) |
            Q(codigo__icontains=query)
        )

        # Redireccionar si hay una única coincidencia
        if productos_list.count() == 1:
            producto = productos_list.first()
            return redirect('detalle_producto', producto_id=producto.pk) # Reemplaza 'detalle_producto' con la URL de tu detalle de producto

    paginator = Paginator(productos_list, 23)

    page = request.GET.get('page', 1)
    try:
        productos = paginator.page(page)
    except EmptyPage:
        productos = paginator.page(paginator.num_pages)

    return render(request, 'productos.html', {'productos': productos, 'paginator': paginator})

@login_required
def sucursales(request):
    sucursales = Sucursal.objects.all()

    return render(request, 'sucursales.html', {'sucursales': sucursales})

@login_required
def agregar_sede(request):
    if request.method == 'POST':
        form = SedeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('../sucursales')
    else:
        form = SedeForm()

    return render(request, 'agregar_sede.html', {'form': form})

@login_required
def detalle_producto(request, producto_id):
    producto = Producto.objects.get(id=producto_id)
    descripcion_limite = obtener_descripcion_limite(producto.descripcion, producto.marca, producto.desc_sub_categoria)
    
    bcv = BCV.objects.order_by('-fecha').first()
    precio_bcv = bcv.precio if bcv else None

    context = {
        'producto': producto,
        'descripcion_limite': descripcion_limite,
        'precio_bcv': precio_bcv,
    }
    
    return render(request, 'producto.html', context)


@login_required
def crear_combo(request):
    if request.method == 'POST':
        codigo_producto = request.POST.get('codigo_producto')
        descripcion = request.POST.get('descripcion')
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_expiracion_str = request.POST.get('fecha_expiracion')
        valor = request.POST.get('valor')
        sede = request.POST.get('sede')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_expiracion = datetime.strptime(fecha_expiracion_str, '%Y-%m-%d').date()

            if fecha_inicio > fecha_expiracion:
                raise ValidationError('La fecha de inicio debe ser anterior a la fecha de expiración.')

            # Obtén el producto asociado al código de barras
            producto = Producto.objects.get(barra=codigo_producto)

            # Crea un nuevo combo asociado al producto
            combo = Combo.objects.create(
                codigo_producto=producto,
                descripcion=descripcion,
                fecha_inicio=fecha_inicio,
                fecha_expiracion=fecha_expiracion,
                valor=valor,
                sede=sede
            )

            # Redirige a la vista deseada después de guardar exitosamente
            return redirect('../')

        except (ValueError, ValidationError, Producto.DoesNotExist) as e:
            # Maneja los errores de formato de fecha, validación o producto no encontrado
            error_message = str(e)
            return render(request, 'agregar_combo.html', {'error_message': error_message})

    return render(request, 'agregar_combo.html')

def ofertasp(request):
   # ofertas = Oferta.objects.all()
   # context = {'ofertas': ofertas}
    
    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[:40] 
    #return render(request, 'ofertas.html', context)
    return render(request, 'ofertafar.html', {'ofertas': ofertas})

def ofertapos(request):
   # ofertas = Oferta.objects.all()
   # context = {'ofertas': ofertas}
    
    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[:10] 
    #return render(request, 'ofertas.html', context)
    return render(request, 'ofertapos.html', {'ofertas': ofertas})

def ofertast(request):
   # ofertas = Oferta.objects.all()
   # context = {'ofertas': ofertas}
    
    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[:40] 
    #return render(request, 'ofertas.html', context)
    return render(request, 'ofertator.html', {'ofertas': ofertas})

def ofertasf(request):
    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[:40]

    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    
    return render(request, 'ofertafar.html', {'ofertas': ofertas})

def ofertasf0(request):
    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[:6]

    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    
    return render(request, 'ofertafar.html', {'ofertas': ofertas})


def ofertasf1(request):
    start_position = 7  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 6  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertasf2(request):
    start_position = 13  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 6  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

from django.shortcuts import render

def promo(request):
    return render(request, 'promo.html')

def ofertast04_1(request):
    start_position = 3  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 10  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast02_1(request):
    start_position = 3  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 10  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_1(request):
    start_position = 3  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 10  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_2(request):
    start_position = 11  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 9  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_3(request):
    start_position = 3  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 10  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_4(request):
    start_position = 11  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 9  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_5(request):
    start_position = 3  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 10  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_6(request):
    start_position = 11  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 9  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_7(request):
    start_position = 3  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 10  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def ofertast16_8(request):
    start_position = 11  # Punto de partida

    if 'start' in request.GET:
        start_position = int(request.GET['start'])

    end_position = start_position + 9  # Punto de finalización

    ofertas = Oferta.objects.filter(linea="MEDICAMENTOS")[start_position:end_position]
    if not ofertas:
        print("No hay ofertas disponibles")
        return redirect('promo') 
    return render(request, 'ofertafar.html', {'ofertas': ofertas, 'start_position': start_position})

def upload_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('productos')
    else:
        form = ProductoForm()
    return render(request, '../productos', {'form': form})